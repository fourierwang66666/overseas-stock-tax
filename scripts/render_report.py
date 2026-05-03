"""
输出三份 Excel：
1. 01_逐笔交易底稿_{年份}.xlsx — 每笔事件 + 4 算法成本基础 + 双汇率轨道
2. 02_B表填报字段_{年份}.xlsx — 直接对应电子税务局 B 表 + 境外抵免明细表
3. 03_四算法对比表_{年份}.xlsx — 四种算法税基差异

用法：
  python3 render_report.py --trades trades.json --year 2024 --out-dir ./output
"""
import argparse
import json
import sys
from decimal import Decimal
from pathlib import Path
from typing import List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("需要 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from cost_basis import Trade, weighted_avg, fifo, moving_avg, specific_id, ALGORITHMS
from compute_tax import compute_capital_gains, market_to_country
from fx_rate import convert
from reconcile import load_trades, D


HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER


def render_ledger(trades: List[Trade], year: int, out: Path):
    """01 逐笔底稿（标准字段 + 4 算法对照）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{year}-逐笔底稿'

    headers = [
        '日期', '账户', '市场', '证券代码', '证券名称', '方向',
        '数量', '成交单价(原币)', '成交金额(原币)', '货币',
        '佣金', '平台费', '印花税', '结算费', 'SEC费', '其他费用', '费用合计',
        # 4 算法
        '加权平均-变动前持仓', '加权平均-变动前成本', '加权平均-变动后持仓', '加权平均-变动后成本',
        '加权平均-平均成本', '加权平均-实际减持收益(原币)',
        'FIFO-实际减持收益(原币)',
        '移动加权-实际减持收益(原币)',
        '个别计价-实际减持收益(原币)',
        # 汇率
        '汇率轨道A(年末)', '加权平均收益(CNY-A)',
        '汇率轨道B(月末锁定)', '加权平均收益(CNY-B)',
    ]
    ws.append(headers)
    _style_header(ws)

    # 跑四算法
    wa = {id(e.trade): e for e in weighted_avg(trades)}
    fi = {id(e.trade): e for e in fifo(trades)}
    ma = {id(e.trade): e for e in moving_avg(trades)}
    si = {id(e.trade): e for e in specific_id(trades)}

    for t in trades:
        if int(t.trade_date[:4]) != year:
            continue
        ew = wa[id(t)]
        ef = fi[id(t)]
        em = ma[id(t)]
        es = si[id(t)]

        cny_a = Decimal(0)
        cny_b = Decimal(0)
        cny_a_disp = '—'
        cny_b_disp = '—'
        if t.side == 'SELL':
            try:
                cny_a = convert(ew.realized_pnl, t.currency, t.trade_date,
                                track='A', settlement_year=year)
                cny_a_disp = float(cny_a)
            except ValueError as e:
                cny_a_disp = f'缺数据: {e}'
            try:
                cny_b = convert(ew.realized_pnl, t.currency, t.trade_date,
                                track='B', settlement_year=year)
                cny_b_disp = float(cny_b)
            except ValueError as e:
                cny_b_disp = f'缺数据: {e}'

        ws.append([
            t.trade_date, t.account, t.market, t.symbol, t.name, t.side,
            float(t.qty), float(t.price), float(t.qty * t.price), t.currency,
            float(t.commission), float(t.platform_fee), float(t.stamp_duty),
            float(t.settlement_fee), float(t.sec_fee), float(t.other_fee),
            float(t.total_fee),
            float(ew.qty_before), float(ew.cost_before),
            float(ew.qty_after), float(ew.cost_after),
            float(ew.avg_cost_before), float(ew.realized_pnl),
            float(ef.realized_pnl),
            float(em.realized_pnl),
            float(es.realized_pnl),
            'A口径', cny_a_disp,
            'B口径', cny_b_disp,
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    wb.save(out)


def render_b_form(trades: List[Trade], year: int, out: Path):
    """02 B 表填报字段（默认加权平均 + 轨道 A 法规口径）"""
    events = weighted_avg(trades)
    cg = compute_capital_gains(events, settlement_year=year, fx_track='A')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'B表-{year}'

    ws.append(['【B 表填报字段速查】（默认: 滚动加权平均 + 法规口径年末汇率）'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(['项目', '金额（CNY）', '说明'])
    _style_header(ws, row=3)

    for key, pool in cg['by_pool'].items():
        ws.append([
            f"{pool['country']} 财产转让所得 - 应纳税所得额",
            float(pool['realized_sum_cny']),
            f"{pool['sell_count']} 笔卖出（同年度同国别互抵后；负数=0）",
        ])
        ws.append([
            f"{pool['country']} 财产转让所得 - 应纳税额",
            float(pool['tax']),
            "= max(应纳税所得额, 0) × 20%",
        ])

    ws.append([])
    ws.append(['合计应纳税额（CNY）', float(cg['total_tax']), '所有国别加总'])

    ws.append([])
    ws.append(['【填报指引】'])
    ws.append(['填报路径', '自然人电子税务局 → 综合所得年度汇算 → 取得境外所得适用 → B 表'])
    ws.append(['填报表格', '《个人所得税年度自行纳税申报表（B 表）（居民个人取得境外所得适用）》'])
    ws.append(['配套表格', '《境外所得个人所得税抵免明细表》'])
    ws.append(['申报时间', f'{year + 1} 年 3 月 1 日 — 6 月 30 日'])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 35
    wb.save(out)


def render_method_compare(trades: List[Trade], year: int, out: Path):
    """03 四算法对比表（含双汇率轨道）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'四算法对比-{year}'

    ws.append([f'四种成本算法 × 两种汇率口径 对比 - {year} 纳税年度'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    headers = ['算法', '汇率轨道', '应纳税所得额合计(CNY)', '应纳税额(CNY) @20%', '与默认差额']
    ws.append(headers)
    _style_header(ws, row=3)

    rows = []
    baseline_tax = None
    for method_name, fn in ALGORITHMS.items():
        events = fn(trades)
        for track, label in [('A', '法规-年末'), ('B', '实操-月末锁定')]:
            try:
                res = compute_capital_gains(events, settlement_year=year, fx_track=track)
                if baseline_tax is None:
                    baseline_tax = res['total_tax']
                rows.append({
                    'method': method_name, 'track': label,
                    'gross': float(res['total_gross_pnl']),
                    'tax': float(res['total_tax']),
                    'diff': float(res['total_tax'] - baseline_tax) if baseline_tax is not None else 0.0,
                })
            except ValueError as e:
                rows.append({
                    'method': method_name, 'track': label,
                    'gross': f'缺数据: {e}', 'tax': '—', 'diff': '—',
                })

    for r in rows:
        ws.append([r['method'], r['track'], r['gross'], r['tax'], r['diff']])

    ws.append([])
    ws.append(['【口径风险提示】'])
    ws.append(['加权平均法', '默认（金标准 / 案例校验版地方税局已接受）'])
    ws.append(['FIFO', 'CRS 国际惯例；某些地方税局可能要求'])
    ws.append(['移动加权', '与加权平均在逐事件粒度等价'])
    ws.append(['个别计价', '需用户在交易时打 lot_tag；事后无法构造'])
    ws.append(['法规-A 口径', '严格符合实施条例第 32 条；100% 被税局接受'])
    ws.append(['实操-B 口径', '主流税务师推荐；避免汇兑收益重复入税基；可能被部分税局拒绝'])
    ws.append([])
    ws.append(['⚠️ 提醒：请就最终口径选择与主管税务机关确认。'])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    wb.save(out)


def render_summary_report(trades: List[Trade], year: int, out: Path):
    """04 税务综合报告 + 合法节税建议 + 强制合规声明"""
    events = weighted_avg(trades)
    cg_a = compute_capital_gains(events, settlement_year=year, fx_track='A')

    sells_count = sum(1 for e in events if e.trade.side == 'SELL')
    buys_count = sum(1 for e in events if e.trade.side == 'BUY')
    has_loss_pool = any(p['realized_sum_cny'] < 0 for p in cg_a['by_pool'].values())
    has_profit_pool = any(p['realized_sum_cny'] > 0 for p in cg_a['by_pool'].values())

    wb = openpyxl.Workbook()

    # ========== Sheet 1: 强制合规声明（首页醒目）==========
    ws = wb.active
    ws.title = '⚠ 重要声明'

    notices = [
        ('【重要声明 — 请务必阅读】', True, 16, 'C00000'),
        ('', False, 11, '000000'),
        ('1. 本报告由 AI 工具自动生成，仅供参考，不构成税务咨询意见或最终申报依据。',
         True, 12, '000000'),
        ('   实际申报金额、算法选择、合理税费扣除范围以主管税务机关认定为准。',
         False, 11, '595959'),
        ('', False, 11, '000000'),
        ('2. 依法纳税是中华人民共和国公民的基本义务（《宪法》第 56 条）。',
         True, 12, '000000'),
        ('   隐匿境外账户、虚构成本、伪造亏损等行为，可能触发《税收征管法》'
         '第 64 条（0.5—5 倍罚款）及《刑法》第 201 条（偷税罪）。',
         False, 11, '595959'),
        ('', False, 11, '000000'),
        ('3. 强烈建议在最终申报前，委托具有执业资格的税务师 / 注册会计师对本报告复核签字，'
         '或直接咨询主管税务机关 12366。',
         True, 12, '000000'),
        ('', False, 11, '000000'),
        ('4. 本 skill 不承担因使用本报告而产生的任何税务、法律或财务责任。',
         False, 11, '595959'),
        ('', False, 11, '000000'),
        ('5. CRS 信息交换已落地，富途/长桥（香港）账户数据已被国税总局获取。'
         '主动自查 + 30 天内补缴，罚款通常可减免；滞纳金（年化 18.25%）不可免。',
         True, 12, 'C00000'),
    ]
    for text, bold, size, color in notices:
        ws.append([text])
        cell = ws.cell(ws.max_row, 1)
        cell.font = Font(bold=bold, size=size, color=color)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 100
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 30

    # ========== Sheet 2: 税务摘要 ==========
    ws2 = wb.create_sheet('税务摘要')
    ws2.append([f'{year} 纳税年度 - 海外股票财产转让所得 综合摘要'])
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.append([])

    ws2.append(['项目', '数值', '说明'])
    _style_header(ws2, row=3)

    summary_rows = [
        ('纳税年度', year, '取得所得年度'),
        ('申报期', f'{year+1}-03-01 至 {year+1}-06-30',
         '逾期触发滞纳金（次年 7-1 起按日万分之五）'),
        ('成本算法', '滚动加权平均法（默认 / 金标准）',
         '另含 FIFO / 移动加权 / 个别计价 三种对比，详见 03 表'),
        ('汇率口径', '法规口径（年末 12-31 中间价）',
         '另含实操口径（买入日锁定原值），详见 03 表'),
        ('卖出笔数', sells_count, ''),
        ('买入笔数', buys_count, ''),
        ('应纳税所得额合计 (CNY)', float(cg_a['total_gross_pnl']),
         '按 (年度, 国别) 分池 → 池内互抵 → 负数池税额=0'),
        ('应纳税额合计 (CNY)', float(cg_a['total_tax']),
         '正数池 × 20%；股息抵免另算'),
    ]
    for r in summary_rows:
        ws2.append(r)

    ws2.append([])
    ws2.append(['【按年度+国别分池明细】'])
    ws2.append(['池子', '国别', '应纳税所得额(CNY)', '应纳税额(CNY)', '卖出笔数'])
    _style_header(ws2, row=ws2.max_row)
    for key, p in cg_a['by_pool'].items():
        ws2.append([
            key, p['country'], float(p['realized_sum_cny']),
            float(p['tax']), p['sell_count'],
        ])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 30

    # ========== Sheet 3: 合法节税建议 ==========
    ws3 = wb.create_sheet('合法节税建议')
    ws3.append(['合法节税建议（仅在中国税法允许的范围内）'])
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.append(['⚠ 严禁：隐匿境外账户、虚构亏损、用他人身份持仓、推迟申报、伪造完税凭证'])
    ws3['A2'].font = Font(bold=True, color='C00000')
    ws3.append([])

    ws3.append(['# ', '建议', '依据 / 适用场景'])
    _style_header(ws3, row=4)

    suggestions = [
        ('1', '充分利用同年度盈亏互抵',
         '国税函[2006]1200 号 — 同一国别同一年度内财产转让所得允许盈亏相抵。'
         '若年内有亏损股票准备清仓，可在同年度卖出已盈利股票时同步处理，降低应纳税基。'),
        ('2', '完整收集合理税费凭证',
         '《个税法实施条例》第 16 条 — 转让收入扣除合理税费。'
         '可扣项：佣金、平台费、印花税(港)、SEC fee(美)、CCASS/DTC 结算费、交易征费。'
         '不扣项：融资利息、汇兑损益。漏扣 1% 费用，税基可能多缴 0.2%。'),
        ('3', '按规定享受境外税收抵免',
         '财政部 税务总局 2020 年第 3 号公告 — 境外已扣税款可凭完税凭证抵免（限额=收入×20%）。'
         '美股 W-8BEN 后预扣 10%，应保留 1042-S；港股股息凭单同理保留。'
         '抵免限额 5 年内可结转。'),
        ('4', '关注沪深港通暂免政策',
         '财政部 税务总局 证监会 2023 年第 23 号公告 — 沪深港通投资港股转让差价暂免至 2027-12-31。'
         '若有新增港股投资计划，可考虑通过沪深港通通道（券商=内地券商）而非富途/长桥（无暂免）。'),
        ('5', '合理选择申报时点',
         '次年 3-1 至 6-30 自行申报。在此期间内任何时间申报均合法，无滞纳金。'
         '6-30 后才申报将按日加收万分之五（年化 18.25%）。'),
        ('6', '主动自查 → 罚款减免',
         '《税收征管法》第 64 条 + 实务共识 — 30 天内主动补税申报，罚款（0.5—5 倍）通常可减免。'
         '滞纳金不可免。被税局约谈后才补缴，处罚更重。'),
        ('7', '长期持仓 vs 短期高频对税负的影响',
         '高频交易 → 频繁触发卖出 → 频繁实现收益 → 当年税基大；'
         '长期持仓 → 浮盈不计税 → 卖出年度才计算。'
         '注意：中国税法对持有时长无差异化税率，但加权平均法下，'
         '长期摊薄成本可能比 FIFO 在涨势中减少税基。'),
        ('8', '股息红利与转让所得分流申报',
         '财政部 税务总局 2020 年第 3 号公告 — 股息按"利息股息红利所得"独立 20%；'
         '不与转让所得互抵。亏损年度仍需就股息补税（2025 补税潮最大盲区）。'),
        ('9', '保留完整凭证至少 5 年',
         '《税收征管法》追溯期一般 3 年，特殊 5 年；境外抵免可结转 5 年。'
         '券商月结单 PDF、完税凭证、汇率工作底稿全部保留。'),
        ('10', '跨年度规划：避免亏损被白白浪费',
         '跨年度亏损不可结转。若大量持仓正处于浮亏状态，可考虑在年内择机实现亏损以匹配当年盈利，'
         '而不是把亏损股票拖到次年（次年若无盈利可冲抵，亏损"作废"）。'
         '⚠ 仅在符合自身投资逻辑时操作，不要为节税而盲目卖股。'),
    ]
    for r in suggestions:
        ws3.append(r)

    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 80
    for r in range(5, ws3.max_row + 1):
        ws3.row_dimensions[r].height = 60
        for c in range(1, 4):
            ws3.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')

    # ========== Sheet 4: 风险提示 ==========
    ws4 = wb.create_sheet('风险提示')
    ws4.append(['风险提示'])
    ws4['A1'].font = Font(bold=True, size=14)
    ws4.append([])
    ws4.append(['# ', '风险', '应对'])
    _style_header(ws4, row=3)

    risks = [
        ('1', 'CRS 信息交换已落地：富途/长桥香港主体已上报中国税局',
         '主动自查申报，避免被约谈'),
        ('2', '亏损年度仍需就股息补税（2025 补税潮最大盲区）',
         '即使资本利得为亏损，股息部分按 20% 单独申报'),
        ('3', '跨年度亏损不可结转',
         '年度规划时考虑在亏损年度择机实现盈利冲抵'),
        ('4', '跨国别盈亏不可互抵（实务主流口径）',
         '美股盈利不能用港股亏损冲抵'),
        ('5', '加权平均法 vs FIFO 选择 — 各地税局口径不一',
         '与主管税局确认；保留 4 算法对比表（03 表）'),
        ('6', '汇率换算双轨争议',
         '与税务师确认采用法规口径还是实操口径'),
        ('7', '富途 2025-06 起关闭内地新户开通',
         '若用富途老账户继续交易，需保持账户合规'),
        ('8', '滞纳金年化 18.25%，多年累积可能超本金',
         '尽早补缴；多年合算可能比逐年算贵'),
    ]
    for r in risks:
        ws4.append(r)
    ws4.column_dimensions['A'].width = 6
    ws4.column_dimensions['B'].width = 50
    ws4.column_dimensions['C'].width = 50
    for r in range(4, ws4.max_row + 1):
        ws4.row_dimensions[r].height = 40
        for c in range(1, 4):
            ws4.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')

    # 智能告警
    if has_loss_pool and has_profit_pool:
        ws4.append([])
        ws4.append(['⚠ 智能检测：本年度同时存在盈利池和亏损池，'
                    '请确认税务师已按"同年度同国别"规则正确互抵'])
        ws4.cell(ws4.max_row, 1).font = Font(bold=True, color='C00000')

    # ========== Sheet 5: 文末重申声明 ==========
    ws5 = wb.create_sheet('再次声明')
    final = [
        '【再次重申】',
        '',
        '✓ 本报告仅供参考，最终申报金额以主管税务机关认定为准',
        '✓ 依法纳税是中国公民的基本义务（《宪法》第 56 条）',
        '✓ 强烈建议委托执业税务师 / 注册会计师复核签字',
        '✓ 本 skill 不承担任何税务、法律或财务责任',
        '✓ 任何涉及隐瞒、虚构、伪造的"避税"建议都是违法行为，本 skill 拒绝提供',
        '',
        '如对申报有疑问，请咨询：',
        '  - 主管税务机关：12366 纳税服务热线',
        '  - 自然人电子税务局：https://etax.chinatax.gov.cn',
        '  - 委托具有执业资格的税务师事务所',
    ]
    for i, line in enumerate(final, 1):
        ws5.append([line])
        if i == 1:
            ws5.cell(i, 1).font = Font(bold=True, size=14, color='C00000')
        elif line.startswith('✓'):
            ws5.cell(i, 1).font = Font(bold=True, size=11)
    ws5.column_dimensions['A'].width = 80

    wb.save(out)


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--trades', required=True)
    p.add_argument('--year', type=int, required=True)
    p.add_argument('--out-dir', default='./output')
    args = p.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    trades = load_trades(args.trades)

    f1 = out_dir / f'01_逐笔交易底稿_{args.year}.xlsx'
    f2 = out_dir / f'02_B表填报字段_{args.year}.xlsx'
    f3 = out_dir / f'03_四算法对比表_{args.year}.xlsx'
    f4 = out_dir / f'04_税务综合报告与合规建议_{args.year}.xlsx'

    render_ledger(trades, args.year, f1)
    render_b_form(trades, args.year, f2)
    render_method_compare(trades, args.year, f3)
    render_summary_report(trades, args.year, f4)

    print(f"✅ 输出四份文件：\n  {f1}\n  {f2}\n  {f3}\n  {f4}")
    print()
    print("⚠ 重要声明：本报告仅供参考，实际申报以主管税务机关认定为准。")
    print("  依法纳税是中国公民的基本义务（《宪法》第 56 条）。")
    print("  建议在最终申报前委托执业税务师 / 注册会计师复核签字。")


if __name__ == '__main__':
    main()
