"""
富途月结单 PDF / Excel 解析器

依赖：pdfplumber、pypdf、openpyxl、pandas

用法：
    python3 parse_futu.py <input.pdf|input.xlsx> [--password XXXX] [--out trades.json]
"""
import argparse
import json
import re
import sys
from decimal import Decimal
from pathlib import Path
from typing import List

try:
    import pdfplumber  # type: ignore
except ImportError:
    pdfplumber = None

try:
    import openpyxl  # type: ignore
except ImportError:
    openpyxl = None

sys.path.insert(0, str(Path(__file__).parent))
from cost_basis import Trade
from schema import Dividend, Transfer


def _to_decimal(s):
    if s is None or s == '':
        return Decimal(0)
    if isinstance(s, (int, float, Decimal)):
        return Decimal(str(s))
    s = str(s).replace(',', '').replace(' ', '').strip()
    if not s or s == '-':
        return Decimal(0)
    return Decimal(s)


def parse_futu_pdf(path: str, password: str = '') -> List[Trade]:
    """
    解析富途月结单 PDF。基于 pdfplumber 文本层提取，禁止 OCR / 视觉识别。
    若 PDF 无文本层（扫描件），抛出 ValueError 提示用户。
    """
    if pdfplumber is None:
        raise RuntimeError("需要安装 pdfplumber：pip install pdfplumber")

    trades: List[Trade] = []
    with pdfplumber.open(path, password=password) as pdf:
        # 检测是否有文本层
        first_page_text = pdf.pages[0].extract_text() if pdf.pages else ''
        if not first_page_text or len(first_page_text.strip()) < 20:
            raise ValueError(
                f"{path} 看起来是扫描件（无文本层）。"
                "skill 禁止使用视觉识别处理 PDF。"
                "请用 OCR 工具（如 tesseract / ocrmypdf）先转为带文本层 PDF，"
                "或导出 Excel 版月结单。"
            )

        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or not table[0]:
                    continue
                # 富途表头识别（中英文都支持）
                header = [str(c or '').strip() for c in table[0]]
                if not _is_futu_trade_header(header):
                    continue
                col_idx = _build_futu_column_index(header)
                for row in table[1:]:
                    if not row or not any(row):
                        continue
                    try:
                        t = _row_to_trade(row, col_idx, source='futu')
                        if t:
                            trades.append(t)
                    except Exception as e:
                        print(f"[WARN] 富途行解析失败：{row}，原因：{e}", file=sys.stderr)
    return trades


def parse_futu_excel(path: str) -> List[Trade]:
    """解析富途导出的 Excel 流水。"""
    if openpyxl is None:
        raise RuntimeError("需要安装 openpyxl：pip install openpyxl")

    trades: List[Trade] = []
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # 找表头行
        header_row_idx = None
        for i, row in enumerate(rows[:10]):
            header = [str(c or '').strip() for c in row]
            if _is_futu_trade_header(header):
                header_row_idx = i
                break
        if header_row_idx is None:
            continue
        header = [str(c or '').strip() for c in rows[header_row_idx]]
        col_idx = _build_futu_column_index(header)
        for row in rows[header_row_idx + 1:]:
            if not row or not any(row):
                continue
            try:
                t = _row_to_trade(list(row), col_idx, source='futu')
                if t:
                    trades.append(t)
            except Exception as e:
                print(f"[WARN] 富途 Excel 行解析失败：{row}，原因：{e}", file=sys.stderr)
    return trades


# 字段映射：富途字段名 → schema key
FUTU_FIELD_MAP = {
    'trade_date': ['成交日期', '交易日期', 'Trade Date', 'Date'],
    'market': ['市场', 'Market', '交易所'],
    'symbol': ['代码', '股票代码', 'Symbol', 'Code'],
    'name': ['名称', '股票名称', 'Name'],
    'side': ['方向', 'Side', '买卖', '类型'],
    'qty': ['成交数量', '数量', 'Quantity', 'Qty'],
    'price': ['成交单价', '价格', 'Price'],
    'amount_local': ['成交金额', '金额', 'Amount'],
    'currency': ['货币', 'Currency', '币种'],
    'commission': ['佣金', 'Commission'],
    'platform_fee': ['平台费', 'Platform Fee'],
    'stamp_duty': ['印花税', 'Stamp Duty'],
    'settlement_fee': ['结算费', 'Settlement Fee'],
    'sec_fee': ['SEC费', 'SEC Fee'],
    'other_fee': ['交易征费', '征费', 'Levy'],
}


def _is_futu_trade_header(header: List[str]) -> bool:
    flat = ' '.join(header)
    return ('成交' in flat or 'Trade' in flat) and ('数量' in flat or 'Qty' in flat or 'Quantity' in flat)


def _build_futu_column_index(header: List[str]) -> dict:
    idx = {}
    for key, names in FUTU_FIELD_MAP.items():
        for col_i, col in enumerate(header):
            if any(n in col for n in names):
                idx[key] = col_i
                break
    return idx


def _row_to_trade(row, col_idx: dict, source: str) -> Trade:
    def get(key, default=''):
        i = col_idx.get(key)
        if i is None or i >= len(row):
            return default
        return row[i]

    side_raw = str(get('side', '')).upper().strip()
    if 'BUY' in side_raw or '买' in side_raw or 'B' == side_raw:
        side = 'BUY'
    elif 'SELL' in side_raw or '卖' in side_raw or 'S' == side_raw:
        side = 'SELL'
    else:
        return None

    qty = _to_decimal(get('qty'))
    price = _to_decimal(get('price'))
    if qty == 0 or price == 0:
        return None

    symbol = str(get('symbol', '')).strip()
    market_raw = str(get('market', '')).strip().upper()
    if 'HK' in market_raw or '港' in market_raw:
        market = 'HK'
        symbol = symbol.zfill(5)  # 港股 5 位前导 0
    elif 'US' in market_raw or '美' in market_raw:
        market = 'US'
    else:
        market = market_raw or 'UNKNOWN'

    return Trade(
        trade_date=str(get('trade_date', ''))[:10],
        market=market,
        symbol=symbol,
        name=str(get('name', '')).strip(),
        side=side,
        qty=qty,
        price=price,
        amount_local=_to_decimal(get('amount_local')) or qty * price,
        currency=str(get('currency', 'HKD')).strip().upper(),
        commission=_to_decimal(get('commission')),
        platform_fee=_to_decimal(get('platform_fee')),
        stamp_duty=_to_decimal(get('stamp_duty')),
        settlement_fee=_to_decimal(get('settlement_fee')),
        sec_fee=_to_decimal(get('sec_fee')),
        other_fee=_to_decimal(get('other_fee')),
        account=f'futu_{source}',
    )


def parse_futu_annual_excel(path: str):
    """
    解析富途【年度账单】Excel（与月结单 Excel 字段不同）。

    年度账单 sheet 结构（11 个 sheet）：
      证券-交易流水    → Trade
      证券-资金进出    → Dividend（备注含 F/D / Cash Dividend / Dividend Tax）
      证券-资产进出    → Transfer（Account Upgrade / 内部账户转移）

    返回 dict：{'trades': [...], 'dividends': [...], 'transfers': [...]}
    """
    if openpyxl is None:
        raise RuntimeError("需要安装 openpyxl：pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    out = {'trades': [], 'dividends': [], 'transfers': []}

    if '证券-交易流水' in wb.sheetnames:
        out['trades'] = _parse_annual_trades(wb['证券-交易流水'])
    if '证券-资金进出' in wb.sheetnames:
        out['dividends'] = _parse_annual_dividends(wb['证券-资金进出'])
    if '证券-资产进出' in wb.sheetnames:
        out['transfers'] = _parse_annual_transfers(wb['证券-资产进出'])

    return out


def _parse_annual_trades(ws):
    """
    年度账单 证券-交易流水 字段：
      成交时间, 账户名称, 账户号码, 品类, 代码名称, 交易所/市场,
      方向, 交收日期, 币种, 数量/面值, 价格, 成交金额, 总费用, 变动金额
    """
    trades = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return trades
    header = [str(c or '').strip() for c in rows[0]]

    def col(name_options):
        for i, h in enumerate(header):
            if any(opt == h or opt in h for opt in name_options):
                return i
        return -1

    idx_time = col(['成交时间'])
    idx_acct = col(['账户名称'])
    idx_sym = col(['代码名称'])
    idx_mkt = col(['交易所/市场', '市场'])
    idx_side = col(['方向'])
    idx_ccy = col(['币种'])
    idx_qty = col(['数量/面值', '数量'])
    idx_price = col(['价格'])
    idx_amount = col(['成交金额'])
    idx_fee = col(['总费用'])

    if any(v < 0 for v in [idx_time, idx_sym, idx_mkt, idx_side, idx_qty, idx_price, idx_amount, idx_fee]):
        raise ValueError(f"富途年度账单字段缺失：header={header}")

    for row in rows[1:]:
        if not row or row[idx_time] is None:
            continue

        side_raw = str(row[idx_side] or '').strip()
        if '买' in side_raw:
            side = 'BUY'
        elif '卖' in side_raw or '沽' in side_raw:
            side = 'SELL'
        else:
            continue

        qty = abs(_to_decimal(row[idx_qty]))
        price = _to_decimal(row[idx_price])
        if qty == 0 or price == 0:
            continue

        symbol = str(row[idx_sym] or '').strip()
        market_raw = str(row[idx_mkt] or '').strip().upper()
        if market_raw in ('SEHK', 'HK') or '港' in market_raw:
            market = 'HK'
            symbol = symbol.zfill(5)
        elif market_raw in ('US', 'NASDAQ', 'NYSE') or '美' in market_raw:
            market = 'US'
        else:
            market = market_raw or 'UNKNOWN'

        trade_date = str(row[idx_time])[:10]  # "2024-01-03 04:52:59" → "2024-01-03"
        currency = str(row[idx_ccy] or '').strip().upper() or ('HKD' if market == 'HK' else 'USD')

        amount = abs(_to_decimal(row[idx_amount]))
        fee = abs(_to_decimal(row[idx_fee]))
        account = str(row[idx_acct] or '').strip() if idx_acct >= 0 else 'futu'

        # 富途年度账单"总费用"为汇总值，明细不可拆分 → 全部归类到 commission
        # 后续若需要分项可基于 market 做粗略推导，但实际税务计算用 total_fee 即可
        trades.append(Trade(
            trade_date=trade_date,
            market=market,
            symbol=symbol,
            name='',  # 年度账单代码名称已合并，不单独 split
            side=side,
            qty=qty,
            price=price,
            amount_local=amount or qty * price,
            currency=currency,
            commission=fee,  # 已含佣金/平台费/印花税/征费/SEC fee 等
            platform_fee=Decimal(0),
            stamp_duty=Decimal(0),
            settlement_fee=Decimal(0),
            sec_fee=Decimal(0),
            other_fee=Decimal(0),
            account='futu',  # 跨账户用同一池（持仓平移已通过 Transfer 处理）
        ))
    return trades


def _parse_annual_dividends(ws):
    """
    年度账单 证券-资金进出 字段：
      日期, 账户名称, 账户号码, 类型, 方向, 币种, 变动金额, 备注
    分红识别：备注含 F/D / DIVIDEND / SCRIP DIVIDEND / 派息 / Cash Dividend
    预扣税识别：备注含 W/H Tax / Withholding / 预扣
    """
    dividends = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return dividends
    header = [str(c or '').strip() for c in rows[0]]

    def col(opts):
        for i, h in enumerate(header):
            if any(o == h or o in h for o in opts):
                return i
        return -1

    idx_date = col(['日期'])
    idx_acct = col(['账户名称'])
    idx_dir = col(['方向'])
    idx_ccy = col(['币种'])
    idx_amt = col(['变动金额'])
    idx_note = col(['备注'])

    for row in rows[1:]:
        if not row or row[idx_date] is None or idx_note < 0:
            continue
        note = str(row[idx_note] or '').strip()
        note_upper = note.upper()
        is_dividend = any(kw in note_upper for kw in ['F/D', 'DIVIDEND', '派息', 'CASH DIV', 'STOCK DIV'])
        is_withhold = any(kw in note_upper for kw in ['W/H TAX', 'WITHHOLDING', '预扣'])
        if not (is_dividend or is_withhold):
            continue
        amt = _to_decimal(row[idx_amt])
        ccy = str(row[idx_ccy] or '').strip().upper()
        date_raw = str(row[idx_date])
        # parse YYYYMMDD → YYYY-MM-DD
        if len(date_raw) == 8 and date_raw.isdigit():
            ex_date = f"{date_raw[:4]}-{date_raw[4:6]}-{date_raw[6:8]}"
        else:
            ex_date = date_raw[:10]

        # 备注里通常含股票代码/名称片段，但年度账单不结构化；保守留空，render 时可由用户补
        gross = amt if is_dividend and amt > 0 else Decimal(0)
        wh = abs(amt) if is_withhold else Decimal(0)

        dividends.append(Dividend(
            ex_date=ex_date,
            pay_date=ex_date,
            market='HK' if ccy == 'HKD' else ('US' if ccy == 'USD' else 'UNKNOWN'),
            symbol='',
            name=note[:40],
            gross_amount=gross,
            withheld_tax=wh,
            currency=ccy,
            account=str(row[idx_acct] or '') if idx_acct >= 0 else 'futu',
        ))
    return dividends


def _parse_annual_transfers(ws):
    """
    年度账单 证券-资产进出 字段：
      日期, 账户名称, 账户号码, 品类, 代码名称, 交易所/市场,
      方向, 类型, 币种, 数量, 备注
    Account Upgrade 等内部转移 → Transfer（不计入 Trade）
    """
    transfers = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return transfers
    header = [str(c or '').strip() for c in rows[0]]

    def col(opts):
        for i, h in enumerate(header):
            if any(o == h or o in h for o in opts):
                return i
        return -1

    idx_date = col(['日期'])
    idx_acct = col(['账户名称'])
    idx_sym = col(['代码名称'])
    idx_mkt = col(['交易所/市场', '市场'])
    idx_dir = col(['方向'])
    idx_qty = col(['数量'])
    idx_note = col(['备注'])

    for row in rows[1:]:
        if not row or row[idx_date] is None:
            continue
        symbol = str(row[idx_sym] or '').strip()
        if not symbol:
            continue
        market_raw = str(row[idx_mkt] or '').strip().upper()
        if market_raw in ('SEHK', 'HK'):
            market = 'HK'
            symbol = symbol.zfill(5)
        elif market_raw == 'US':
            market = 'US'
        else:
            market = market_raw or 'UNKNOWN'
        direction = str(row[idx_dir] or '').strip()
        qty = _to_decimal(row[idx_qty])
        date_raw = str(row[idx_date])
        if len(date_raw) == 8 and date_raw.isdigit():
            tdate = f"{date_raw[:4]}-{date_raw[4:6]}-{date_raw[6:8]}"
        else:
            tdate = date_raw[:10]
        transfers.append(Transfer(
            transfer_date=tdate,
            market=market,
            symbol=symbol,
            name='',
            qty=qty,
            direction=direction,
            from_account=str(row[idx_acct] or '') if direction.lower() in ('out',) else '',
            to_account=str(row[idx_acct] or '') if direction.lower() in ('in',) else '',
            note=str(row[idx_note] or '')[:60] if idx_note >= 0 else '',
        ))
    return transfers


def main():
    p = argparse.ArgumentParser(description='富途月结单 / 年度账单解析器')
    p.add_argument('input', help='富途 PDF 或 Excel 路径')
    p.add_argument('--password', default='', help='PDF 密码')
    p.add_argument('--annual', action='store_true', help='Excel 是富途年度账单格式（多 sheet）')
    p.add_argument('--out', default='-', help='输出 JSON 路径，- 为 stdout')
    args = p.parse_args()

    ext = Path(args.input).suffix.lower()
    if ext == '.pdf':
        trades = parse_futu_pdf(args.input, args.password)
        result = {'trades': trades, 'dividends': [], 'transfers': []}
    elif ext in ('.xlsx', '.xls'):
        if args.annual:
            result = parse_futu_annual_excel(args.input)
        else:
            result = {'trades': parse_futu_excel(args.input), 'dividends': [], 'transfers': []}
        trades = result['trades']
    else:
        print(f"❌ 不支持的格式：{ext}", file=sys.stderr)
        sys.exit(1)

    print(f"[OK] 解析 {args.input}：{len(trades)} 笔交易", file=sys.stderr)
    out = [
        {**t.__dict__, 'qty': str(t.qty), 'price': str(t.price),
         'amount_local': str(t.amount_local),
         'commission': str(t.commission), 'platform_fee': str(t.platform_fee),
         'stamp_duty': str(t.stamp_duty), 'settlement_fee': str(t.settlement_fee),
         'sec_fee': str(t.sec_fee), 'other_fee': str(t.other_fee)}
        for t in trades
    ]
    if args.out == '-':
        print(json.dumps(out, ensure_ascii=False, indent=2))
    else:
        Path(args.out).write_text(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
