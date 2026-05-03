"""
浏览器版 Excel 解析器 — 不依赖 pdfplumber/pypdf
支持：
- 富途月结单 Excel / 年度账单 Excel
- 长桥月结单 Excel
- 通用 schema Excel（用户按模板整理）

返回统一的 Trade 列表（含 schema.Dividend / Transfer / CrossBrokerTransfer / CorporateAction）。
"""
import sys
from decimal import Decimal
from pathlib import Path
from typing import List

import openpyxl

from cost_basis import Trade


def _D(v):
    if v is None or v == '':
        return Decimal(0)
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).replace(',', '').replace(' ', '').strip()
    if not s or s == '-':
        return Decimal(0)
    try:
        return Decimal(s)
    except Exception:
        return Decimal(0)


# 通用字段别名（覆盖富途、长桥、自定义模板）
HEADER_ALIASES = {
    'trade_date': ['成交日期', '交易日期', '交收日', 'Trade Date', 'Date', '成交时间'],
    'market': ['市场', 'Market', '交易所', '交易所/市场'],
    'symbol': ['代码', '股票代码', 'Symbol', 'Code', '代码名称'],
    'name': ['名称', '股票名称', 'Name'],
    'side': ['方向', 'Side', '买卖', '类型'],
    'qty': ['成交数量', '数量', '数量/面值', 'Quantity', 'Qty'],
    'price': ['成交单价', '价格', 'Price', '成交价'],
    'amount_local': ['成交金额', '金额', 'Amount'],
    'currency': ['货币', 'Currency', '币种'],
    'commission': ['佣金', 'Commission'],
    'platform_fee': ['平台费', 'Platform Fee', '平台使用费'],
    'stamp_duty': ['印花税', 'Stamp Duty'],
    'settlement_fee': ['结算费', 'Settlement Fee', 'CCASS'],
    'sec_fee': ['SEC费', 'SEC Fee', 'FINRA'],
    'other_fee': ['交易征费', '征费', 'Levy', '总费用'],
}


def _is_trade_header(header: List[str]) -> bool:
    flat = ' '.join(header)
    has_date = any(k in flat for k in ['日期', 'Date', '时间'])
    has_qty = any(k in flat for k in ['数量', 'Qty', 'Quantity'])
    has_side = any(k in flat for k in ['方向', '买卖', 'Side', '类型'])
    return has_date and has_qty and has_side


def _build_idx(header: List[str]) -> dict:
    idx = {}
    for key, names in HEADER_ALIASES.items():
        for i, col in enumerate(header):
            col_str = str(col or '').strip()
            if any(n in col_str for n in names):
                idx[key] = i
                break
    return idx


def _row_to_trade(row, col_idx: dict, account_hint: str) -> Trade:
    def get(key, default=''):
        i = col_idx.get(key)
        if i is None or i >= len(row):
            return default
        return row[i]

    side_raw = str(get('side', '')).upper().strip()
    if any(k in side_raw for k in ['BUY', '买']) or side_raw == 'B':
        side = 'BUY'
    elif any(k in side_raw for k in ['SELL', '卖', '沽']) or side_raw == 'S':
        side = 'SELL'
    else:
        return None

    qty = abs(_D(get('qty')))
    price = _D(get('price'))
    if qty == 0 or price == 0:
        return None

    symbol = str(get('symbol', '')).strip()
    market_raw = str(get('market', '')).strip().upper()
    if 'HK' in market_raw or '港' in market_raw or 'SEHK' in market_raw:
        market = 'HK'
        symbol = symbol.zfill(5)
    elif 'US' in market_raw or '美' in market_raw or 'NYSE' in market_raw or 'NASDAQ' in market_raw:
        market = 'US'
    elif 'SG' in market_raw or '新加坡' in market_raw:
        market = 'SG'
    else:
        market = market_raw or 'UNKNOWN'

    raw_date = str(get('trade_date', ''))
    trade_date = raw_date[:10] if len(raw_date) >= 10 else raw_date

    return Trade(
        trade_date=trade_date,
        market=market,
        symbol=symbol,
        name=str(get('name', '')).strip(),
        side=side,
        qty=qty,
        price=price,
        amount_local=_D(get('amount_local')) or qty * price,
        currency=str(get('currency', 'HKD')).strip().upper(),
        commission=_D(get('commission')),
        platform_fee=_D(get('platform_fee')),
        stamp_duty=_D(get('stamp_duty')),
        settlement_fee=_D(get('settlement_fee')),
        sec_fee=_D(get('sec_fee')),
        other_fee=_D(get('other_fee')),
        account=account_hint,
    )


def parse_any_excel(path: str) -> List[Trade]:
    """自动识别 sheet 中的交易表头并解析"""
    if not path.lower().endswith(('.xlsx', '.xls', '.csv')):
        raise ValueError(f"暂不支持 {path}（浏览器版仅支持 Excel/CSV，PDF 请切到 AI 模式或先转为 Excel）")

    if path.lower().endswith('.csv'):
        return _parse_csv(path)

    trades: List[Trade] = []
    wb = openpyxl.load_workbook(path, data_only=True)

    # 账户名提示：从文件名猜
    fn = path.lower()
    if 'futu' in fn or '富途' in fn or 'moomoo' in fn:
        account = 'futu'
    elif 'longbridge' in fn or 'longport' in fn or '长桥' in fn:
        account = 'longbridge'
    else:
        account = 'unknown'

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # 找表头行（前 15 行扫描）
        header_row_idx = None
        for i, row in enumerate(rows[:15]):
            header = [str(c or '').strip() for c in row]
            if _is_trade_header(header):
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        header = [str(c or '').strip() for c in rows[header_row_idx]]
        col_idx = _build_idx(header)
        if 'qty' not in col_idx or 'price' not in col_idx or 'side' not in col_idx:
            continue

        for row in rows[header_row_idx + 1:]:
            if not row or not any(row):
                continue
            try:
                t = _row_to_trade(list(row), col_idx, account)
                if t:
                    trades.append(t)
            except Exception as e:
                print(f"[WARN] {sheet_name} 行解析失败：{e}", file=sys.stderr)

    return trades


def _parse_csv(path: str) -> List[Trade]:
    import csv
    trades: List[Trade] = []
    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return trades
    header = [str(c or '').strip() for c in rows[0]]
    col_idx = _build_idx(header)
    if 'qty' not in col_idx:
        return trades
    fn = path.lower()
    account = 'futu' if 'futu' in fn or '富途' in fn else 'longbridge' if '长桥' in fn else 'unknown'
    for row in rows[1:]:
        if not row or not any(row):
            continue
        try:
            t = _row_to_trade(row, col_idx, account)
            if t:
                trades.append(t)
        except Exception as e:
            print(f"[WARN] CSV 行错: {e}", file=sys.stderr)
    return trades
